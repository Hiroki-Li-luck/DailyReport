import openpyxl
import glob
from datetime import datetime, timedelta

today = datetime.now()
weekday = today.weekday()
w_list = ['(月)', '(火)', '(水)', '(木)', '(金)', '(土)', '(日)']
dt = today.strftime('%m' + '/' + '%d')
wd = w_list[weekday]

f = open('日報.txt', 'w')
f.write(dt + wd + '作業内容を下記に記載いたします。　\n\n ')

report_name = None
target_file = glob.glob("作業_*.xlsx")
print(target_file)
wb = openpyxl.load_workbook(target_file[0], data_only=True)

# ️全てのシートをタプル化
worksheets = ('ST4602', 'ST4301', 'ST2404')
testers = ("C3:C100")
report_name = None

# シートの取得をループ
for worksheet in worksheets:
    ws = wb[worksheet]

    # シート名をテキストファイルに記載
    with open('日報.txt', 'a') as f:
        print('◆' + worksheet, file=f)
    f.close()
    no_tester_flag = 0

    # 列取得のループ
    for tester in ws[testers]:

        # セルの取得のループ
        for tester_cell in tester:
            tester_name = tester_cell.value

            # セルに名前が入っているかどうかで条件分岐
            if tester_name != None:
                no_tester_flag = 1
                target_row = int(tester_cell.row)
                report_name = ws.cell(row=target_row, column=2).value

                # 検証仕様書名をテキストファイルに記載
                with open('日報.txt', 'a') as f:
                    print(report_name, file=f)
                f.close()
                target_count = tester_cell.offset(0, 1).value

                # 連想配列作成
                report_name = {tester_name: target_count}
                next_tester = tester_cell.offset(0, 2)
                next_tester_name = next_tester.value

                if report_name == {}:
                    with open('日報.txt', 'a') as f:
                        print('-\n', file=f)
                    f.close()

                # 次のテスト者の名前があるかどうかで条件分岐
                else:
                    while next_tester_name != None:
                        target_count = next_tester.offset(0, 1).value
                        report_name[next_tester_name] = target_count
                        next_tester = next_tester.offset(0, 2)
                        next_tester_name = next_tester.value

                    # 連想配列からテスト実施者と件数を取得
                    for name in report_name:

                        # テキストファイルにテスト実施者と件数を記載
                        with open('日報.txt', 'a') as f:
                            print(
                                '　　　' + '実施件数:' + ' ' + str(report_name[name]) + '件' + '(' + name + ')', file=f)
                        f.close()

                with open('日報.txt', 'a') as f:
                    print('', file=f)
                f.close()

    if no_tester_flag == 0:
        with open('日報.txt', 'a') as f:
            print('-\n', file=f)
        f.close()

with open('日報.txt', 'a') as f:
    f.write(
        'その他\nなし\n\n■再現確認\nなし\n\n■デグレ確認\nなし\n\n■改修確認\nなし\n\n■不具合起票\nなし\n\n以上、よろしくお願い致します。')
f.close()
